import json
import os
import urllib.error
import urllib.parse
import urllib.request

import openpyxl

from config import API_TOKEN, BASE_URL_MISSION, BASE_URL_USERS


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ARQUIVO_MATRIZ = os.path.join(BASE_DIR, "input", "matriz_treinamentos.xlsx")
STATUS_VALIDOS = {"N/A", "A", "B", "T", "A.T", "B.T"}


def _json_request(url, headers=None, params=None, method="GET", payload=None, timeout=40):
    if params:
        query = urllib.parse.urlencode(params)
        url = f"{url}{'&' if '?' in url else '?'}{query}"
    data = None
    req_headers = dict(headers or {})
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
        req_headers.setdefault("Content-Type", "application/json")
    req = urllib.request.Request(url, data=data, headers=req_headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            body = resp.read()
            return resp.status, json.loads(body.decode("utf-8", errors="replace") or "{}")
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {e.code}: {body[:500]}")


def _buscar_user_id(matricula):
    status, dados = _json_request(
        f"{BASE_URL_USERS}/v2/users",
        headers={"Content-Type": "application/json", "Authorization": API_TOKEN},
        params={"username__eq": matricula, "find_exact_match": "true", "limit": 1},
        timeout=40,
    )
    lista = dados.get("users", dados.get("results", dados.get("data", [])))
    if not lista:
        raise RuntimeError(f"Matricula {matricula} nao encontrada na Skore.")
    user_id = str(lista[0].get("id", "")).strip()
    if not user_id:
        raise RuntimeError(f"Usuario da matricula {matricula} veio sem ID.")
    return user_id


def _associar_missao(matricula, mission_id):
    user_id = _buscar_user_id(matricula)
    _json_request(
        f"{BASE_URL_MISSION}/missions/add_audience",
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {API_TOKEN}"},
        method="PATCH",
        payload={"mission_id": mission_id, "user_ids": [user_id], "team_ids": []},
        timeout=40,
    )
    return user_id


def atualizar_status_matriz(matricula, mission_id, status):
    matricula = str(matricula or "").strip()
    mission_id = str(mission_id or "").strip()
    status = str(status or "").strip().upper()
    if status in {"NA", "N.A", "NÃO", "NAO"}:
        status = "N/A"
    if status not in STATUS_VALIDOS:
        raise ValueError("Status invalido. Use N/A, A, B, T, A.T ou B.T.")
    if not matricula or not mission_id:
        raise ValueError("Matricula e missao sao obrigatorias.")

    user_id = ""
    associado = False
    if status != "N/A":
        user_id = _associar_missao(matricula, mission_id)
        associado = True

    wb = openpyxl.load_workbook(ARQUIVO_MATRIZ)
    ws = wb["Matriz"] if "Matriz" in wb.sheetnames else wb.active
    headers = {str(cell.value).strip().lower(): idx for idx, cell in enumerate(ws[1], 1)}
    col_mat = headers.get("matricula")
    col_mid = headers.get("id da missao") or headers.get("mission_id")
    col_status = headers.get("status matriz")
    if not col_mat or not col_mid or not col_status:
        raise RuntimeError("A matriz precisa das colunas Matricula, ID da missao e Status Matriz.")

    linha_alvo = None
    for row in range(2, ws.max_row + 1):
        mat = str(ws.cell(row, col_mat).value or "").strip()
        mid = str(ws.cell(row, col_mid).value or "").strip()
        if mat == matricula and mid == mission_id:
            linha_alvo = row
            break
    if not linha_alvo:
        raise RuntimeError("Combinacao matricula x missao nao encontrada na matriz.")

    anterior = str(ws.cell(linha_alvo, col_status).value or "").strip() or "N/A"
    ws.cell(linha_alvo, col_status).value = status
    wb.save(ARQUIVO_MATRIZ)
    return {"matricula": matricula, "mission_id": mission_id, "status": status, "anterior": anterior, "associado": associado, "user_id": user_id}
