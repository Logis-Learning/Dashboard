"""
Skore — Extrator de Status de Matrículas
Login automático + barra de progresso + requisições em lote.
 
Instalar:
    pip install playwright openpyxl
    playwright install chromium
"""
 
import json, csv, sys, os
from datetime import datetime
from playwright.sync_api import sync_playwright

# Garante UTF-8 no output (necessário no Windows)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
 
# ============================================================
# CONFIGURAÇÃO — edite aqui
# ============================================================
SKORE_URL  = "https://universidadesimpar.skore.io"
EMAIL      = "felipe.pianelli@jsl.com.br"
SENHA      = "199003Fg?"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ARQUIVO_MISSOES = os.path.join(BASE_DIR, "input", "missoes_para_exportar.xlsx")
ARQUIVO_ATIVOS  = os.path.join(BASE_DIR, "input", "Ativos.xlsx")
 
LIMIT       = 5000   # limite por lote
LOTE_SIZE   = 200    # matrículas por requisição
# ============================================================
 
import openpyxl
 
def ler_planilha(arquivo, nome_coluna, filtro_status=False):
    """
    Lê coluna de uma planilha Excel.
    Se filtro_status=True, só retorna linhas onde coluna Status = 'Ativo' ou 'A'.
    """
    if not arquivo or not os.path.exists(arquivo):
        if arquivo:
            print(f"⚠️  '{arquivo}' não encontrada — ignorando filtro")
        return []
    try:
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in ws[1]]
 
        idx = next((i for i, h in enumerate(headers) if nome_coluna.lower() in h.lower()), None)
        if idx is None:
            print(f"⚠️  Coluna '{nome_coluna}' não encontrada. Disponíveis: {headers}")
            return []
 
        # Índice da coluna Status (coluna R = índice 17)
        idx_status = None
        if filtro_status:
            idx_status = 17  # coluna R fixa
            print(f"   Usando coluna R (índice 17) como Status. Cabeçalho: '{headers[17] if len(headers) > 17 else 'fora do range'}'")
 
        valores = []
        ignorados = 0
        for r in range(2, ws.max_row + 1):
            valor = ws.cell(row=r, column=idx+1).value
            if not valor:
                continue
 
            # Aplica filtro de status se solicitado
            if filtro_status and idx_status is not None:
                status_val = str(ws.cell(row=r, column=idx_status+1).value or "").strip().upper()
                if status_val not in ("ATIVO", "A"):
                    ignorados += 1
                    continue
 
            valores.append(str(valor).strip())
 
        if filtro_status:
            print(f"   ✅ {len(valores)} ativos | ⏭️  {ignorados} ignorados (status diferente de Ativo/A)")
 
        return valores
    except Exception as e:
        print(f"⚠️  Erro ao ler '{arquivo}': {e}")
        return []
 
missoes      = ler_planilha(ARQUIVO_MISSOES, "mission_id")
ativos       = ler_planilha(ARQUIVO_ATIVOS,  "Número Pessoal", filtro_status=True)
MISSION_ID   = ",".join(missoes)
 
LOOKER_URL = "https://skoreio.sa.looker.com"
 
fields = ",".join([
    "users.id", "users.name", "users.email", "users.username",
    "mission_definitions.mission_id", "mission_definitions.name",
    "mission_definitions.workload_minutes", "mission_definitions.is_mandatory",
    "mission_definitions.allow_enrollment_cancellation",
    "enrollments.detailed_status", "enrollments.average_progress",
    "enrollments.enrolled_time", "enrollments.completed_time",
    "enrollments.due_date_available_at_date",
    "enrollments.cancelled_reason", "enrollments.cancelled_at_date",
    "certificates.created_date",
    "metadata.cpf", "metadata.cargo", "metadata.data_admissao",
    "metadata.departamento", "metadata.filial",
    "leaders.name_list", "departments.mensure",
])
 
def montar_url(usernames=""):
    return (
        f"{LOOKER_URL}/explore/skore/mission_definitions.json"
        f"?fields={fields}"
        f"&f[users.isDeleted]=No"
        f"&f[users.active]=Yes"
        f"&f[departments.deleted_date]=NULL"
        f"&f[enrollments.detailed_status]="
        f"&f[mission_definitions.mission_id]={MISSION_ID}"
        f"&f[users.username]={usernames}"
        f"&sorts=users.name"
        f"&limit={LIMIT}"
    )
 
def progresso(passo, total, label, largura=40):
    preenchido = int(largura * passo / total)
    barra = "█" * preenchido + "░" * (largura - preenchido)
    pct = int(100 * passo / total)
    sys.stdout.write(f"\r  [{barra}] {pct:3d}%  {label}")
    sys.stdout.flush()
    if passo == total:
        print()
 
print("\n🚀 Skore — Extrator de Status de Matrículas")
print("=" * 50)
if missoes:
    print(f"📋 {len(missoes)} missões carregadas")
else:
    print("📋 Buscando todas as missões")
if ativos:
    print(f"👤 {len(ativos)} usuários carregados — serão enviados em lotes de {LOTE_SIZE}")
else:
    print("👤 Buscando todos os usuários")
 
# Divide ativos em lotes
if ativos:
    lotes = [ativos[i:i+LOTE_SIZE] for i in range(0, len(ativos), LOTE_SIZE)]
else:
    lotes = [[]]  # um lote vazio = busca todos
 
TOTAL_ETAPAS = 3 + len(lotes) + 2  # login + looker + lotes + processar + salvar
etapa_atual  = [0]
 
def avancar(label):
    etapa_atual[0] += 1
    progresso(etapa_atual[0], TOTAL_ETAPAS, label)
 
all_rows = []
 
with sync_playwright() as p:
 
    avancar("Iniciando browser...")
    browser = p.chromium.launch(headless=True)
    context = browser.new_context()
    page    = context.new_page()
 
    avancar("Fazendo login...")
    page.goto(f"{SKORE_URL}/login", wait_until="networkidle")
    page.fill("#username", EMAIL)
    page.fill("#password", SENHA)
    page.wait_for_timeout(2000)
    page.wait_for_selector("#login-button:not([disabled])", timeout=10000)
    page.click("#login-button")
    page.wait_for_timeout(3000)
 
    if "/login" in page.url:
        print("\n❌ Falha no login — verifique email e senha")
        browser.close()
        sys.exit(1)
 
    avancar("Carregando sessão Looker...")
    # Navega para o relatório da Skore que contém o Looker embed
    page.goto(f"{SKORE_URL}/reports/5955", wait_until="networkidle", timeout=90000)
    page.wait_for_timeout(10000)
 
    # Agora navega para o domínio do Looker mantendo a sessão
    page.goto(f"{LOOKER_URL}/embed/dashboards/2122", wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(5000)
 
    for i, lote in enumerate(lotes):
        label = f"Baixando lote {i+1}/{len(lotes)}..."
        avancar(label)
 
        url = montar_url(",".join(lote))
 
        texto = page.evaluate("""
            async (url) => {
                const resp = await fetch(url, {
                    method: "GET",
                    credentials: "include",
                    headers: {
                        "Accept": "application/json",
                        "x-csrf-token": document.cookie.match(/CSRF-TOKEN=([^;]+)/)?.[1] || ""
                    }
                });
                return await resp.text();
            }
        """, url)
 
        if not texto.strip() or texto.strip() == "[]":
            print(f"\n   ⚠️  Lote {i+1} vazio")
            continue
 
        try:
            dados = json.loads(texto)
            all_rows.extend(dados)
        except Exception as e:
            print(f"\n   ❌ Erro no lote {i+1}: {texto[:100]}")
 
    browser.close()
 
avancar("Processando...")
 
if not all_rows:
    print("\n❌ Nenhum dado retornado")
    sys.exit(1)
 
def get_status(row):
    return str(row.get("enrollments.detailed_status") or "").upper()
 
not_started = [r for r in all_rows if get_status(r) == "NOT_STARTED"]
in_progress  = [r for r in all_rows if get_status(r) == "IN_PROGRESS"]
completed    = [r for r in all_rows if get_status(r) in ("COMPLETED", "COMPLETED_AFTER_DUE_DATE")]
expired      = [r for r in all_rows if get_status(r) == "EXPIRED"]
 
avancar("Salvando resultados...")

import pandas as pd

OUTPUT_DIR    = os.path.join(BASE_DIR, "output_powerbi")
os.makedirs(OUTPUT_DIR, exist_ok=True)

filename      = os.path.join(OUTPUT_DIR, "skore_resultado.csv")
filename_json = os.path.join(OUTPUT_DIR, "skore_resultado.json")
filename_xlsx = os.path.join(OUTPUT_DIR, "relatorio_matriculas.xlsx")

# CSV e JSON brutos
fieldnames = list(all_rows[0].keys())
with open(filename, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(all_rows)

with open(filename_json, "w", encoding="utf-8") as f:
    json.dump(all_rows, f, ensure_ascii=False, indent=2)

# Converte para relatorio_matriculas.xlsx (formato que o dashboard lê)
MAPA = {
    "users.id":                                          "ID do Usuário",
    "users.name":                                        "Nome",
    "users.email":                                       "Email",
    "users.username":                                    "Username",
    "metadata.cpf":                                      "CPF",
    "metadata.cargo":                                    "Cargo",
    "metadata.departamento":                             "Departamento",
    "metadata.filial":                                   "Filial",
    "leaders.name_list":                                 "Lista de Nome dos Lideres",
    "metadata.data_admissao":                            "Data de admissao",
    "mission_definitions.mission_id":                    "ID da missao",
    "mission_definitions.name":                          "Missao",
    "mission_definitions.is_mandatory":                  "Missao Obrigatoria?",
    "mission_definitions.allow_enrollment_cancellation": "Missao pode ser cancelada?",
    "mission_definitions.workload_minutes":              "Carga horaria (min)",
    "enrollments.enrolled_time":                         "Data da Matricula",
    "enrollments.due_date_available_at_date":            "Prazo de Conclusao",
    "enrollments.detailed_status":                       "Status Detalhado da Matricula",
    "enrollments.completed_time":                        "Data de Conclusao da Matricula",
    "enrollments.cancelled_at_date":                     "Data de Cancelamento da Matricula",
    "enrollments.cancelled_reason":                      "Motivo do Cancelamento da matricula",
    "certificates.created_date":                         "Data de Emissao do Certificado",
    "enrollments.average_progress":                      "Progresso",
    "departments.mensure":                               "Times do usuario",
}
df = pd.DataFrame(all_rows).rename(columns=MAPA)
for col in ["Setor", "Turno", "Media das notas", "Tempo de treinamento (min)"]:
    if col not in df.columns:
        df[col] = ""
df.to_excel(filename_xlsx, index=False)

print("\n" + "=" * 50)
print(f"✅ {len(all_rows):,} registros extraídos")
print(f"   📭 Not Started  : {len(not_started):,}")
print(f"   🔄 In Progress  : {len(in_progress):,}")
print(f"   ✅ Completed    : {len(completed):,}")
print(f"   ⏰ Expired      : {len(expired):,}")
print(f"\n📁 {filename_xlsx}")
print("🏁 Concluído!")
